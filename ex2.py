import tkinter as tk
from tkinter import ttk
from datetime import datetime
from openpyxl import load_workbook

filename=r'C:\Users\mizun\OneDrive\MyPythonProject\出退勤ツール.xlsx'

def openWorkbook():
    wb=load_workbook(filename)
    ws=wb.active
    return (wb,ws)

def closeWorkbook(wb):
    wb.save(filename)
    variable.set('')

def begin():
    now=datetime.now()
    wb,ws=openWorkbook()
    ws.append([now.date(),now.time(),variable.get()])
    button_begin['state']=tk.DISABLED
    button_finish['state']=tk.NORMAL
    closeWorkbook(wb)

def finish():
    now=datetime.now()
    wb,ws=openWorkbook()
    max_row=ws.max_row
    ws.cell(max_row,4).value=now.time()
    ws.cell(max_row,5).value=variable.get()
    button_begin['state']=tk.NORMAL
    button_finish['state']=tk.DISABLED
    wb.save(r'C:\Users\mizun\OneDrive\MyPythonProject\出退勤ツール.xlsx')
    closeWorkbook(wb)
  

root=tk.Tk()
root.title('TITLE')

button_begin=ttk.Button(text='出勤',command=begin)
button_begin.pack(side='left')

button_finish=ttk.Button(text='退勤',command=finish,state=tk.DISABLED)
button_finish.pack(side='left')

variable=tk.StringVar()
entry=ttk.Entry(textvariable=variable)
entry.pack(side='left')

root.mainloop()